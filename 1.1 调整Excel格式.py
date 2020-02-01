# coding:utf-8
# author:stay5sec

import openpyxl

# desk path：/Users/super/Desktop/

# 读取复制数据
wb1 = openpyxl.load_workbook("/Users/super/Desktop/一份假数据.xlsx")

# 读取模板
wb2 = openpyxl.load_workbook("/Users/super/Desktop/模板/模板.xlsx")

sheets1 = wb1.sheetnames
sheets2 = wb2.sheetnames

print("\n原始数据表格有：", sheets1)
print("模板中的表格有：", sheets2)

# exit()

sheet1 = wb1[sheets1[0]]
sheet2 = wb2[sheets2[0]]

max_row = sheet1.max_row  # 行
max_col = sheet1.max_column  # 列

print("\n原始数据\n行数量：{}, 列数量:{}".format(max_row, max_col))

# exit()

for m in range(1, max_row + 1):
    for n in range(97, 97 + max_col):
        sheet2[chr(n) + str(m)] = sheet1[chr(n) + str(m)].value

wb2.save("/Users/super/Desktop/整理过的假数据.xlsx")

wb1.close()
wb2.close()
