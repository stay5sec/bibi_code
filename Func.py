# coding:utf-8
# author:stay5sec

import openpyxl

def excel_beautiful(or_name,temp_sheet,tar_name):

    # 读取复制数据
    wb1 = openpyxl.load_workbook("/Users/super/Desktop/{}.xlsx".format(or_name))

    # 读取模板
    wb2 = openpyxl.load_workbook("/Users/super/Desktop/模板/模板.xlsx")

    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[temp_sheet]]

    max_row = sheet1.max_row  # 行
    max_col = sheet1.max_column  # 列

    print("\n原始数据\n行数量：{}, 列数量:{}".format(max_row, max_col))

    for m in range(1, max_row + 1):
        for n in range(97, 97 + max_col):
            sheet2[chr(n) + str(m)] = sheet1[chr(n) + str(m)].value

    wb2.save("/Users/super/Desktop/{}.xlsx".format(tar_name))

    wb1.close()
    wb2.close()

    print("\n表格整理完毕！请知悉……")


import shutil
import os

tar_path = "/Users/super/Desktop/books/"

out_path = "/Users/super/Desktop/books2/"

def opath(x):
    return os.path.join("/Users/super/Desktop",x)


def change_file_suffix(tar_path,out_path,types):
    # 判断文件是否存在
    if os.path.exists(out_path):
        print("请查看已存在文件！")
        exit()
    else:
        os.mkdir(out_path)
        print("输出文件夹创建完毕……")

    for f in os.listdir(tar_path):
        f_new = f.split(".")[0] + ".{}".format(types)

        tar_file = os.path.join(tar_path, f)
        out_file = os.path.join(out_path, f_new)

        shutil.copy(tar_file, out_file)

        print("{} - 文件修改完毕".format(f))

